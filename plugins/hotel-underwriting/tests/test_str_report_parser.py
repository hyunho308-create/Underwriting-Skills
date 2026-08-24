from __future__ import annotations

from copy import deepcopy
from datetime import date
import hashlib
import json
from pathlib import Path
import subprocess
import sys

from openpyxl import Workbook, load_workbook

from _test_paths import PLUGIN_ROOT, add_scripts_to_path


add_scripts_to_path()

from contract_store import (  # noqa: E402
    build_source_record,
    initialize_project,
    read_jsonl,
    write_jsonl_atomic,
)
from str_report_parser import parse_str_workbook  # noqa: E402


MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def month_add(month: date, offset: int) -> date:
    total = month.year * 12 + month.month - 1 + offset
    return date(total // 12, total % 12 + 1, 1)


def add_response(workbook: Workbook, suffix: str, subject: str) -> None:
    sheet = workbook.create_sheet("Response" + suffix)
    sheet["B1"] = "Tab 22 - Response Report - Performance Set" if not suffix else "Tab 37 - Response Report - Comp Set #2"
    sheet["B2"] = f"{subject}        1 Main St        Example, CA 90000"
    sheet["B3"] = "STR ID: 100        Operator: Example"
    sheet["B4"] = "For the Month of: March 2026        Date Created: April 17, 2026"
    for cell, value in (("C22", "STR#"), ("D22", "Name"), ("H22", "Rooms"), ("L22", "Open Date")):
        sheet[cell] = value
    sheet["C23"], sheet["D23"], sheet["H23"], sheet["L23"] = 100, subject, 250, 198806
    sheet["C24"], sheet["D24"], sheet["H24"], sheet["L24"] = 200 + (2 if suffix else 1), f"Comp Hotel {suffix or '1'}", 180, 200101


def add_comp(workbook: Workbook, suffix: str, subject: str, value_offset: int) -> None:
    sheet = workbook.create_sheet("Comp" + suffix)
    sheet["B1"] = "Tab 4 - Competitive Set Report - Performance Set" if not suffix else "Tab 24 - Competitive Set Report - Comp Set #2"
    sheet["AF1"] = "Currency USD - U.S. Dollar"
    sheet["B2"] = f"{subject}        1 Main St        Example, CA 90000"
    sheet["B4"] = "For the Month of: March 2026        Date Created: April 17, 2026"
    report_month = date(2026, 3, 1)
    months = [month_add(report_month, offset) for offset in range(-17, 1)]
    for title_row, metric, base in ((19, "Occupancy (%)", 70), (31, "ADR", 200), (43, "RevPAR", 140)):
        sheet.cell(title_row, 2, metric)
        sheet.cell(title_row, 22, "Year to Date")
        sheet.cell(title_row, 30, "Running 12 Month")
        for index, month in enumerate(months, start=3):
            sheet.cell(title_row + 1, index, MONTH_NAMES[month.month - 1])
        for index, year in enumerate((2024, 2025, 2026), start=22):
            sheet.cell(title_row + 1, index, year)
        for index, year in enumerate((2024, 2025, 2026), start=30):
            sheet.cell(title_row + 1, index, year)
        sheet.cell(title_row + 2, 2, "My Property")
        sheet.cell(title_row + 3, 2, "Competitive Set")
        sheet.cell(title_row + 4, 2, "Index (MPI)" if metric.startswith("Occupancy") else "Index (ARI)" if metric == "ADR" else "Index (RGI)")
        for column in list(range(3, 21)) + list(range(22, 25)) + list(range(30, 33)):
            prop = base + value_offset + column / 10
            comp = base - 5 + column / 10
            sheet.cell(title_row + 2, column, prop)
            sheet.cell(title_row + 3, column, comp)
            sheet.cell(title_row + 4, column, prop / comp * 100)


def add_day_of_week(workbook: Workbook, suffix: str, value_offset: int) -> None:
    sheet = workbook.create_sheet("Day of Week" + suffix)
    sheet["D15"], sheet["J15"], sheet["P15"] = "Occupancy (%)", "Average Daily Rate", "RevPAR"
    for column, value in ((4, "My Property"), (6, "Competitive Set"), (8, "Index (MPI)"), (10, "My Property"), (12, "Competitive Set"), (14, "Index (ARI)"), (16, "My Property"), (18, "Competitive Set"), (20, "Index (RGI)")):
        sheet.cell(16, column, value)
    for label_row, value_row, label in ((60, 63, "Weekday"), (66, 69, "Weekend")):
        sheet.cell(label_row, 2, label)
        sheet.cell(value_row, 3, "Running 12 Month")
        for prop_col, comp_col, index_col, base in ((4, 6, 8, 70), (10, 12, 14, 220), (16, 18, 20, 150)):
            prop = base + value_offset
            comp = base - 5
            sheet.cell(value_row, prop_col, prop)
            sheet.cell(value_row, comp_col, comp)
            sheet.cell(value_row, index_col, prop / comp * 100)


def add_segmentation(workbook: Workbook, suffix: str, value_offset: int) -> None:
    for base_name, metric_base in (("Segmentation Occ", 30), ("Segmentation ADR", 180), ("Segmentation RevPAR", 60)):
        sheet = workbook.create_sheet(base_name + suffix)
        sheet["D6"] = base_name.replace("Segmentation ", "")
        sheet["Q6"] = "Percent Change (%)"
        for column, segment in ((4, "Transient"), (7, "Group"), (10, "Contract")):
            sheet.cell(7, column, segment)
            sheet.cell(8, column, "My Prop")
            sheet.cell(8, column + 1, "Comp Set")
        for column, segment in ((17, "Transient"), (20, "Group"), (23, "Contract")):
            sheet.cell(7, column, segment)
            sheet.cell(8, column, "My Prop")
            sheet.cell(8, column + 1, "Comp Set")
        sheet["B38"] = "Running 12 Month"
        for row, year in enumerate((2024, 2025, 2026), start=39):
            sheet.cell(row, 2, year)
            for column, segment_offset in ((4, 0), (7, 10), (10, 20)):
                sheet.cell(row, column, metric_base + value_offset + segment_offset + year - 2024)
                sheet.cell(row, column + 1, metric_base - 5 + segment_offset + year - 2024)
            for column in (17, 20, 23):
                sheet.cell(row, column, -10 - row)
                sheet.cell(row, column + 1, -5 - row)


def build_str_workbook(path: Path) -> None:
    workbook = Workbook()
    contents = workbook.active
    contents.title = "Table of Contents"
    contents["B2"] = "STR ID 100 / Fiscal Year January 1 / Weekends: Friday & Saturday"
    contents["B5"] = "Currency: US Dollar / Competitive Set Data Excludes Subject Property"
    for suffix, offset in (("", 0), ("_2", 20)):
        add_response(workbook, suffix, "Example Hotel")
        add_comp(workbook, suffix, "Example Hotel", offset)
        add_day_of_week(workbook, suffix, offset)
        add_segmentation(workbook, suffix, offset)
    workbook.save(path)


def retarget_str_workbook(
    path: Path,
    report_month: date,
    *,
    primary_competitor_str_id: int | None = None,
) -> None:
    workbook = load_workbook(path)
    month_label = report_month.strftime("%B %Y")
    created = month_add(report_month, 1)
    created_label = created.strftime("%B 17, %Y")
    months = [month_add(report_month, offset) for offset in range(-17, 1)]
    for suffix in ("", "_2"):
        workbook["Comp" + suffix]["B4"] = (
            f"For the Month of: {month_label}        Date Created: {created_label}"
        )
        workbook["Response" + suffix]["B4"] = (
            f"For the Month of: {month_label}        Date Created: {created_label}"
        )
        for title_row in (19, 31, 43):
            for column, month in enumerate(months, start=3):
                workbook["Comp" + suffix].cell(title_row + 1, column, MONTH_NAMES[month.month - 1])
    if primary_competitor_str_id is not None:
        workbook["Response"]["C24"] = primary_competitor_str_id
        workbook["Response"]["D24"] = f"Comp Hotel {primary_competitor_str_id}"
    workbook.save(path)


def primitive_by_kind(parsed: dict, kind: str) -> dict:
    return next(
        item["data"]["information_primitive"]
        for item in parsed["evidence"]
        if item["data"]["information_primitive"]["attributes"]["observation_kind"] == kind
    )


def test_parses_two_comp_sets_and_requested_periods(tmp_path: Path) -> None:
    workbook_path = tmp_path / "example-str.xlsx"
    build_str_workbook(workbook_path)

    parsed = parse_str_workbook(workbook_path, "src_0123456789abcdef")

    assert [item["comp_set_label"] for item in parsed] == ["Performance Set", "Comp Set #2"]
    assert all(item["status"] == "complete" for item in parsed)
    monthly = primitive_by_kind(parsed[0], "monthly_performance")["attributes"]["rows"]
    assert primitive_by_kind(parsed[0], "monthly_performance")["attributes"]["currency"] == "USD"
    assert primitive_by_kind(parsed[0], "monthly_performance")["attributes"]["units"]["adr"] == "USD"
    assert len(monthly) == 18
    assert monthly[0]["period"] == "2024-10"
    assert monthly[-1]["period_end"] == "2026-03-31"
    ytd = primitive_by_kind(parsed[0], "ytd_performance")["attributes"]["rows"]
    assert [row["period_end"] for row in ytd] == ["2024-03-31", "2025-03-31", "2026-03-31"]
    running_12 = primitive_by_kind(parsed[0], "running_12_performance")["attributes"]["rows"]
    assert running_12[-1]["period_start"] == "2025-04-01"
    membership = primitive_by_kind(parsed[0], "comp_set_membership")["attributes"]["hotels"]
    assert membership == [
        {"role": "subject", "name": "Example Hotel", "rooms": 250, "str_id": "100", "open_date": "1988-06"},
        {"role": "competitive_hotel", "name": "Comp Hotel 1", "rooms": 180, "str_id": "201", "open_date": "2001-01"},
    ]
    day_type = primitive_by_kind(parsed[0], "day_type_running_12")["attributes"]
    assert day_type["weekend_definition"] == "Friday & Saturday"
    assert {row["day_type"] for row in day_type["rows"]} == {"weekday", "weekend"}
    segmentation = primitive_by_kind(parsed[0], "segment_running_12")["attributes"]["rows"]
    assert len(segmentation) == 9
    assert {row["segment"] for row in segmentation} == {"transient", "group", "contract"}
    first_segment = next(row for row in segmentation if row["year"] == 2024 and row["segment"] == "transient")
    assert first_segment["metrics"]["occupancy"]["property"] == 30
    assert first_segment["metrics"]["adr"]["property"] == 180
    assert first_segment["metrics"]["revpar"]["property"] == 60
    assert primitive_by_kind(parsed[0], "segment_running_12")["attributes"]["value_kind"] == "actual"


def test_extract_str_cli_is_idempotent_and_writes_no_facts(tmp_path: Path) -> None:
    deal_room = tmp_path / "Example Hotel"
    deal_room.mkdir()
    workbook_path = deal_room / "example-str.xlsx"
    build_str_workbook(workbook_path)
    initialize_project(deal_room, "example-hotel", "Example Hotel", primary_property_name="Example Hotel")
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    source = build_source_record(
        "example-hotel",
        "local_file",
        "relative_path",
        "example-str.xlsx",
        content_sha256=digest,
        captured_at="2026-08-19T00:00:00Z",
        document_type="performance.str_report",
        metadata={"extension": ".xlsx"},
    )
    source["availability"] = "available"
    write_jsonl_atomic(deal_room / ".hotel-underwriting" / "sources.jsonl", [source])
    command = [
        sys.executable,
        str(PLUGIN_ROOT / "scripts" / "contracts_cli.py"),
        "extract-str-data",
        "--deal-room",
        str(deal_room),
        "--entity-name",
        "Model Chosen Hotel",
    ]

    first = json.loads(subprocess.run(command, check=True, capture_output=True, text=True).stdout)
    evidence_path = deal_room / ".hotel-underwriting" / "evidence.jsonl"
    first_evidence = read_jsonl(evidence_path)
    stale = deepcopy(
        next(
            record
            for record in first_evidence
            if record["data"]["information_primitive"]["attributes"]["observation_kind"]
            == "segment_running_12"
        )
    )
    stale["evidence_id"] = "ev_0000000000000000"
    stale["locator"] = {"kind": "workbook_range", "value": "'Segmentation Occ'!B38:Z41"}
    write_jsonl_atomic(evidence_path, [*first_evidence, stale])
    second = json.loads(subprocess.run(command, check=True, capture_output=True, text=True).stdout)

    assert first["comp_sets_selected"] == 2
    assert first["evidence_added"] == 12
    assert first["facts_added"] == 0
    assert second["evidence_added"] == -1
    evidence = read_jsonl(evidence_path)
    assert len(evidence) == 12
    assert all(record["evidence_id"] != "ev_0000000000000000" for record in evidence)
    assert {record["data"]["information_primitive"]["type"] for record in evidence} == {"property.str"}
    assert {record["data"]["information_primitive"]["subject"] for record in evidence} == {"Model Chosen Hotel"}
    assert {
        record["data"]["information_primitive"]["attributes"]["source_reported_subject"]
        for record in evidence
    } == {"Example Hotel"}
    assert {
        record["data"]["information_primitive"]["attributes"]["entity_name_resolution"]
        for record in evidence
    } == {"model_supplied"}
    assert not (deal_room / ".hotel-underwriting" / "facts.jsonl").exists()
    coverage = json.loads(
        (deal_room / ".hotel-underwriting" / "derived" / "str-information-evidence-coverage.json").read_text(encoding="utf-8")
    )
    assert coverage["status"] == "complete"
    assert coverage["facts_added"] == 0
    assert coverage["entity_name_resolution"] == "model_supplied"
    assert {item["source_reported_subject"] for item in coverage["comp_sets"]} == {"Example Hotel"}


def test_extract_str_cli_preserves_all_reports_and_membership_versions(tmp_path: Path) -> None:
    deal_room = tmp_path / "Example Hotel"
    deal_room.mkdir()
    earlier_path = deal_room / "str-2026-02.xlsx"
    current_path = deal_room / "str-2026-03.xlsx"
    build_str_workbook(earlier_path)
    build_str_workbook(current_path)
    retarget_str_workbook(
        earlier_path,
        date(2026, 2, 1),
        primary_competitor_str_id=301,
    )
    initialize_project(deal_room, "example-hotel", "Example Hotel", primary_property_name="Example Hotel")

    sources = []
    for captured_at, path in (
        ("2026-03-01T00:00:00Z", earlier_path),
        ("2026-04-01T00:00:00Z", current_path),
    ):
        source = build_source_record(
            "example-hotel",
            "local_file",
            "relative_path",
            path.name,
            content_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
            captured_at=captured_at,
            document_type="performance.str_report",
            metadata={"extension": ".xlsx"},
        )
        source["availability"] = "available"
        sources.append(source)
    write_jsonl_atomic(deal_room / ".hotel-underwriting" / "sources.jsonl", sources)

    command = [
        sys.executable,
        str(PLUGIN_ROOT / "scripts" / "contracts_cli.py"),
        "extract-str-data",
        "--deal-room",
        str(deal_room),
        "--entity-name",
        "Model Chosen Hotel",
    ]
    first = json.loads(subprocess.run(command, check=True, capture_output=True, text=True).stdout)
    second = json.loads(subprocess.run(command, check=True, capture_output=True, text=True).stdout)

    assert first["comp_sets_selected"] == 4
    assert first["evidence_added"] == 24
    assert second["evidence_added"] == 0
    evidence = read_jsonl(deal_room / ".hotel-underwriting" / "evidence.jsonl")
    monthly = [
        record
        for record in evidence
        if record["data"]["information_primitive"]["attributes"]["observation_kind"]
        == "monthly_performance"
    ]
    assert len(monthly) == 4
    monthly_attrs = [record["data"]["information_primitive"]["attributes"] for record in monthly]
    assert {attrs["report_month"] for attrs in monthly_attrs} == {"2026-02", "2026-03"}

    primary_versions = {
        attrs["comp_set_version_id"]
        for attrs in monthly_attrs
        if attrs["comp_set_key"] == "performance-set"
    }
    secondary_versions = {
        attrs["comp_set_version_id"]
        for attrs in monthly_attrs
        if attrs["comp_set_key"] == "comp-set-2"
    }
    assert len(primary_versions) == 2
    assert len(secondary_versions) == 1
    assert {attrs["comp_set_version_basis"] for attrs in monthly_attrs} == {"membership"}

    coverage = json.loads(
        (deal_room / ".hotel-underwriting" / "derived" / "str-information-evidence-coverage.json").read_text(encoding="utf-8")
    )
    assert coverage["selection_policy"] == "all available STR reports"
    assert coverage["older_reports_not_selected"] == []
    assert {
        (item["report_month"], item["monthly_period_start"], item["monthly_period_end"])
        for item in coverage["comp_sets"]
    } == {
        ("2026-02", "2024-09", "2026-02"),
        ("2026-03", "2024-10", "2026-03"),
    }


def test_extract_str_cli_requires_model_entity_name_before_persisting(tmp_path: Path) -> None:
    deal_room = tmp_path / "Example Hotel"
    deal_room.mkdir()
    workbook_path = deal_room / "example-str.xlsx"
    build_str_workbook(workbook_path)
    initialize_project(deal_room, "example-hotel", "Example Hotel")
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    source = build_source_record(
        "example-hotel",
        "local_file",
        "relative_path",
        "example-str.xlsx",
        content_sha256=digest,
        captured_at="2026-08-19T00:00:00Z",
        document_type="performance.str_report",
        metadata={"extension": ".xlsx"},
    )
    source["availability"] = "available"
    write_jsonl_atomic(deal_room / ".hotel-underwriting" / "sources.jsonl", [source])
    result = subprocess.run(
        [
            sys.executable,
            str(PLUGIN_ROOT / "scripts" / "contracts_cli.py"),
            "extract-str-data",
            "--deal-room",
            str(deal_room),
        ],
        capture_output=True,
        text=True,
    )

    assert result.returncode != 0
    assert "--entity-name is required" in result.stderr
    assert not (deal_room / ".hotel-underwriting" / "evidence.jsonl").exists()
