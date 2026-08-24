"""Deterministic extraction of standardized STR/STAR workbook data."""

from __future__ import annotations

import calendar
from collections import defaultdict
from datetime import date, datetime
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Iterable


PRODUCER = {"skill": "extract-str-data", "version": "0.2.0"}
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
REQUIRED_EVIDENCE_KINDS = {
    "comp_set_membership",
    "monthly_performance",
    "ytd_performance",
    "running_12_performance",
    "day_type_running_12",
    "segment_running_12",
}

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


class StrWorkbookError(ValueError):
    """Raised when an STR workbook cannot be interpreted safely."""


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def _comp_set_version_identity(
    logical_label: str,
    subject: str,
    subject_str_id: str | None,
    hotels: list[dict[str, Any]] | None,
    source_id: str,
) -> tuple[str, str]:
    if hotels:
        members = []
        for hotel in hotels:
            if hotel.get("role") != "competitive_hotel":
                continue
            member_id = hotel.get("str_id")
            members.append(f"str:{member_id}" if member_id else f"name:{_norm(hotel.get('name'))}")
        payload = {
            "logical_comp_set": logical_label,
            "subject": f"str:{subject_str_id}" if subject_str_id else f"name:{_norm(subject)}",
            "competitive_members": sorted(set(members)),
        }
        basis = "membership"
    else:
        payload = {
            "logical_comp_set": logical_label,
            "subject": f"str:{subject_str_id}" if subject_str_id else f"name:{_norm(subject)}",
            "source_id": source_id,
        }
        basis = "source_fallback"
    digest = hashlib.sha256(
        json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()[:16]
    return f"csv_{digest}", basis


def _value(row: list[Any], column: int) -> Any:
    return row[column - 1] if 0 < column <= len(row) else None


def _number(value: Any) -> float | int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text or text.casefold() in {"n/a", "na", "nm", "-"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace("$", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        parsed = float(text)
    except ValueError:
        return None
    if negative:
        parsed = -parsed
    return int(parsed) if parsed.is_integer() else parsed


def _integer(value: Any) -> int | None:
    parsed = _number(value)
    if parsed is None or isinstance(parsed, float) and not parsed.is_integer():
        return None
    return int(parsed)


def _string_id(value: Any) -> str | None:
    integer = _integer(value)
    if integer is not None:
        return str(integer)
    text = str(value or "").strip()
    return text or None


def _month_number(value: Any) -> int | None:
    return MONTHS.get(_norm(value))


def _month_add(month: date, offset: int) -> date:
    total = month.year * 12 + month.month - 1 + offset
    return date(total // 12, total % 12 + 1, 1)


def _month_end(month: date) -> date:
    return date(month.year, month.month, calendar.monthrange(month.year, month.month)[1])


def _period_for_month(month: date) -> dict[str, str]:
    return {
        "period": month.strftime("%Y-%m"),
        "period_start": month.isoformat(),
        "period_end": _month_end(month).isoformat(),
    }


def _ytd_period(year: int, report_month: date) -> dict[str, str | int]:
    end = date(year, report_month.month, calendar.monthrange(year, report_month.month)[1])
    return {"year": year, "period_start": date(year, 1, 1).isoformat(), "period_end": end.isoformat()}


def _running_12_period(year: int, report_month: date) -> dict[str, str | int]:
    end_month = date(year, report_month.month, 1)
    start = _month_add(end_month, -11)
    return {
        "year": year,
        "period_start": start.isoformat(),
        "period_end": _month_end(end_month).isoformat(),
    }


def _open_month(value: Any) -> str | None:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m")
    integer = _integer(value)
    if integer is None:
        match = re.search(r"\b(\d{4})[-/]?(\d{2})\b", str(value or ""))
        if not match:
            return None
        integer = int(match.group(1) + match.group(2))
    text = str(integer)
    if len(text) != 6:
        return None
    year, month = int(text[:4]), int(text[4:])
    if year < 1800 or not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"


def _column_name(column: int) -> str:
    result = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _range(sheet: str, min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'!{_column_name(min_col)}{min_row}:{_column_name(max_col)}{max_row}"


def _needed_sheet_name(name: str) -> bool:
    return bool(
        re.fullmatch(
            r"(?:Table of Contents|Comp|Response|Day of Week|Segmentation Occ|Segmentation ADR|Segmentation RevPAR)(?:_\d+)?",
            name,
            re.IGNORECASE,
        )
    )


def _rows_from_xlsx(path: Path) -> dict[str, list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment-specific failure
        raise StrWorkbookError(".xlsx extraction requires the bundled openpyxl package") from exc
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        return {
            sheet.title: [[cell.value for cell in row] for row in sheet.iter_rows()]
            for sheet in workbook.worksheets
            if _needed_sheet_name(sheet.title)
        }
    finally:
        workbook.close()


def _rows_from_xls(path: Path) -> dict[str, list[list[Any]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - optional legacy dependency
        raise StrWorkbookError(
            "legacy .xls extraction requires xlrd; prefer the newest equivalent .xlsx report "
            "or save the workbook as .xlsx without changing its cells"
        ) from exc
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        return {
            sheet.name: [sheet.row_values(row) for row in range(sheet.nrows)]
            for sheet in workbook.sheets()
            if _needed_sheet_name(sheet.name)
        }
    finally:
        workbook.release_resources()


def load_workbook_rows(path: Path) -> dict[str, list[list[Any]]]:
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(path)
    if suffix == ".xls":
        return _rows_from_xls(path)
    raise StrWorkbookError(f"unsupported STR workbook extension: {path.suffix}")


def _all_header_text(rows: list[list[Any]], limit: int = 8) -> str:
    return " ".join(str(value) for row in rows[:limit] for value in row if value not in (None, ""))


def _parse_report_month(rows: list[list[Any]]) -> date:
    match = re.search(
        r"For the Month of:\s*([A-Za-z]+)\s+(\d{4})",
        _all_header_text(rows),
        re.IGNORECASE,
    )
    if not match:
        raise StrWorkbookError("could not find the STR report month")
    month = MONTHS.get(match.group(1).casefold())
    if not month:
        raise StrWorkbookError(f"unrecognized STR report month: {match.group(1)}")
    return date(int(match.group(2)), month, 1)


def _parse_created_date(rows: list[list[Any]]) -> str | None:
    match = re.search(
        r"Date Created:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})",
        _all_header_text(rows),
        re.IGNORECASE,
    )
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%B %d, %Y").date().isoformat()
    except ValueError:
        return None


def _parse_currency(rows: list[list[Any]]) -> str | None:
    text = _all_header_text(rows)
    match = re.search(r"\bCurrency\s*:?\s*([A-Z]{3})\b", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    normalized = _norm(text)
    if "currency us dollar" in normalized or "currency u s dollar" in normalized:
        return "USD"
    return None


def _parse_weekend_definition(sheets: dict[str, list[list[Any]]]) -> str | None:
    for sheet_name, rows in sheets.items():
        if not sheet_name.casefold().startswith("table of contents"):
            continue
        match = re.search(
            r"Weekends:\s*(.*?)(?=\s*/|\s+Currency\s*:|$)",
            _all_header_text(rows, 10),
            re.IGNORECASE,
        )
        if match:
            return match.group(1).strip()
    return None


def _first_text(rows: list[list[Any]], row_number: int) -> str | None:
    if not 1 <= row_number <= len(rows):
        return None
    for value in rows[row_number - 1]:
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _subject_from_header(rows: list[list[Any]]) -> str | None:
    line = _first_text(rows, 2)
    if not line:
        return None
    return re.split(r"\s{4,}", line, maxsplit=1)[0].strip() or None


def _title(rows: list[list[Any]]) -> str:
    return _first_text(rows, 1) or "Competitive Set Report"


def _comp_set_identity(title: str, suffix: str) -> tuple[str, str, int | None]:
    match = re.search(r"Comp(?:etitive)?\s+Set\s*#?\s*(\d+)", title, re.IGNORECASE)
    if match:
        number = int(match.group(1))
        return f"comp-set-{number}", f"Comp Set #{number}", number
    if re.search(r"Performance\s+Set", title, re.IGNORECASE):
        return "performance-set", "Performance Set", None
    if suffix and int(suffix) > 1:
        number = int(suffix)
        return f"comp-set-{number}", f"Comp Set #{number}", number
    return "competitive-set", "Competitive Set", None


def _find_response_hotels(rows: list[list[Any]]) -> tuple[list[dict[str, Any]], str | None]:
    header_row = None
    columns: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=1):
        labels = {_norm(value): column for column, value in enumerate(row, start=1) if value is not None}
        if "name" in labels and "rooms" in labels and "open date" in labels and "str" in labels:
            header_row = row_number
            columns = {
                "str_id": labels["str"],
                "name": labels["name"],
                "rooms": labels["rooms"],
                "open_date": labels["open date"],
            }
            break
    if header_row is None:
        raise StrWorkbookError("Response sheet does not contain the hotel response table")

    hotels: list[dict[str, Any]] = []
    last_row = header_row
    for row_number in range(header_row + 1, len(rows) + 1):
        row = rows[row_number - 1]
        name = str(_value(row, columns["name"]) or "").strip()
        rooms = _integer(_value(row, columns["rooms"]))
        if not name or rooms is None:
            if hotels:
                break
            continue
        hotel: dict[str, Any] = {
            "role": "subject" if not hotels else "competitive_hotel",
            "name": name,
            "rooms": rooms,
        }
        str_id = _string_id(_value(row, columns["str_id"]))
        open_month = _open_month(_value(row, columns["open_date"]))
        if str_id:
            hotel["str_id"] = str_id
        if open_month:
            hotel["open_date"] = open_month
        hotels.append(hotel)
        last_row = row_number

    if not hotels:
        raise StrWorkbookError("Response sheet hotel table is empty")
    locator = _range(
        "Response",
        header_row,
        min(columns.values()),
        last_row,
        max(columns.values()),
    )
    return hotels, locator


def _metric_name(value: Any) -> str | None:
    label = _norm(value)
    if label in {"occupancy", "occupancy percent"}:
        return "occupancy"
    if label in {"adr", "average daily rate"}:
        return "adr"
    if label == "revpar":
        return "revpar"
    return None


def _find_row_label(rows: list[list[Any]], start_row: int, labels: set[str], limit: int = 10) -> int | None:
    for row_number in range(start_row, min(len(rows), start_row + limit - 1) + 1):
        row = rows[row_number - 1]
        for value in row[:6]:
            if _norm(value) in labels:
                return row_number
    return None


def _metric_point(property_value: Any, comp_value: Any, index_value: Any = None) -> dict[str, Any]:
    prop = _number(property_value)
    comp = _number(comp_value)
    index = _number(index_value)
    if index is None and prop is not None and comp not in (None, 0):
        index = prop / comp * 100
    return {"property": prop, "comp_set": comp, "penetration_index": index}


def _validate_actual_metric_point(point: dict[str, Any], metric: str, context: str) -> None:
    for role in ("property", "comp_set"):
        value = point.get(role)
        if value is None:
            continue
        if metric == "occupancy" and not 0 <= value <= 100:
            raise StrWorkbookError(f"{context} has invalid actual occupancy for {role}: {value}")
        if metric in {"adr", "revpar"} and value < 0:
            raise StrWorkbookError(f"{context} has invalid actual {metric} for {role}: {value}")


def _parse_comp(rows: list[list[Any]], sheet_name: str, report_month: date) -> dict[str, Any]:
    sections: dict[str, dict[str, Any]] = {}
    for row_number, row in enumerate(rows, start=1):
        found = None
        for column, raw in enumerate(row[:10], start=1):
            metric = _metric_name(raw)
            if metric:
                found = (metric, column)
                break
        if not found or found[0] in sections:
            continue
        metric, label_column = found
        header_row = row_number + 1
        if header_row > len(rows):
            continue
        period_headers = rows[header_row - 1]

        ytd_column = next(
            (column for column, value in enumerate(row, start=1) if _norm(value) in {"year to date", "ytd"}),
            None,
        )
        running_12_column = next(
            (column for column, value in enumerate(row, start=1) if _norm(value) == "running 12 month"),
            None,
        )
        if ytd_column is None or running_12_column is None:
            continue

        month_columns = [
            column
            for column in range(1, ytd_column)
            if _month_number(_value(period_headers, column)) is not None
        ]
        if len(month_columns) < 18:
            raise StrWorkbookError(f"{sheet_name} {metric} section has fewer than 18 monthly columns")
        month_columns = month_columns[-18:]

        property_row = _find_row_label(rows, row_number + 2, {"my property", "my prop"})
        comp_row = _find_row_label(rows, row_number + 2, {"competitive set", "comp set"})
        index_row = _find_row_label(rows, row_number + 2, {"index mpi", "index ari", "index rgi"})
        if property_row is None or comp_row is None:
            raise StrWorkbookError(f"{sheet_name} {metric} section is missing property or comp-set rows")

        def aggregate_columns(start_column: int) -> list[tuple[int, int]]:
            result: list[tuple[int, int]] = []
            for column in range(start_column, min(len(period_headers), start_column + 6) + 1):
                year = _integer(_value(period_headers, column))
                if year and 1900 <= year <= 2200:
                    result.append((column, year))
                if len(result) == 3:
                    break
            return result

        ytd_columns = aggregate_columns(ytd_column)
        running_12_columns = aggregate_columns(running_12_column)
        if len(ytd_columns) != 3 or len(running_12_columns) != 3:
            raise StrWorkbookError(f"{sheet_name} {metric} section does not contain three YTD and TTM years")

        prop_values = rows[property_row - 1]
        comp_values = rows[comp_row - 1]
        index_values = rows[index_row - 1] if index_row else []
        expected_months = [_month_add(report_month, offset) for offset in range(-17, 1)]
        source_months = [_month_number(_value(period_headers, column)) for column in month_columns]
        if source_months != [month.month for month in expected_months]:
            raise StrWorkbookError(f"{sheet_name} {metric} monthly headers do not end at the report month")

        sections[metric] = {
            "title_row": row_number,
            "last_data_row": max(property_row, comp_row, index_row or 0),
            "month_columns": month_columns,
            "monthly": [
                _metric_point(
                    _value(prop_values, column),
                    _value(comp_values, column),
                    _value(index_values, column),
                )
                for column in month_columns
            ],
            "ytd_columns": ytd_columns,
            "ytd": {
                year: _metric_point(
                    _value(prop_values, column),
                    _value(comp_values, column),
                    _value(index_values, column),
                )
                for column, year in ytd_columns
            },
            "running_12_columns": running_12_columns,
            "running_12": {
                year: _metric_point(
                    _value(prop_values, column),
                    _value(comp_values, column),
                    _value(index_values, column),
                )
                for column, year in running_12_columns
            },
            "label_column": label_column,
        }

    if set(sections) != {"occupancy", "adr", "revpar"}:
        missing = sorted({"occupancy", "adr", "revpar"} - set(sections))
        raise StrWorkbookError(f"{sheet_name} is missing metric sections: {missing}")

    expected_months = [_month_add(report_month, offset) for offset in range(-17, 1)]
    monthly = []
    for index, month in enumerate(expected_months):
        row = _period_for_month(month)
        row["metrics"] = {metric: sections[metric]["monthly"][index] for metric in sections}
        monthly.append(row)

    ytd_years = sorted(set.intersection(*(set(section["ytd"]) for section in sections.values())))
    running_12_years = sorted(
        set.intersection(*(set(section["running_12"]) for section in sections.values()))
    )
    ytd = []
    for year in ytd_years:
        row = _ytd_period(year, report_month)
        row["metrics"] = {metric: sections[metric]["ytd"][year] for metric in sections}
        ytd.append(row)
    running_12 = []
    for year in running_12_years:
        row = _running_12_period(year, report_month)
        row["metrics"] = {metric: sections[metric]["running_12"][year] for metric in sections}
        running_12.append(row)

    first_row = min(section["title_row"] for section in sections.values())
    last_row = max(section["last_data_row"] for section in sections.values())
    monthly_min = min(min(section["month_columns"]) for section in sections.values())
    monthly_max = max(max(section["month_columns"]) for section in sections.values())
    ytd_min = min(min(column for column, _ in section["ytd_columns"]) for section in sections.values())
    ytd_max = max(max(column for column, _ in section["ytd_columns"]) for section in sections.values())
    ttm_min = min(
        min(column for column, _ in section["running_12_columns"]) for section in sections.values()
    )
    ttm_max = max(
        max(column for column, _ in section["running_12_columns"]) for section in sections.values()
    )
    return {
        "monthly": monthly,
        "ytd": ytd,
        "running_12": running_12,
        "locators": {
            "monthly": _range(sheet_name, first_row, monthly_min, last_row, monthly_max),
            "ytd": _range(sheet_name, first_row, ytd_min, last_row, ytd_max),
            "running_12": _range(sheet_name, first_row, ttm_min, last_row, ttm_max),
        },
    }


def _metric_columns(rows: list[list[Any]], max_rows: int = 30) -> dict[str, dict[str, int]]:
    for row_number, row in enumerate(rows[:max_rows], start=1):
        starts = [(column, _metric_name(value)) for column, value in enumerate(row, start=1)]
        starts = [(column, metric) for column, metric in starts if metric]
        if {metric for _, metric in starts} != {"occupancy", "adr", "revpar"}:
            continue
        starts.sort()
        labels_row = rows[row_number] if row_number < len(rows) else []
        result: dict[str, dict[str, int]] = {}
        for index, (start, metric) in enumerate(starts):
            end = starts[index + 1][0] - 1 if index + 1 < len(starts) else min(len(labels_row), start + 8)
            columns: dict[str, int] = {}
            for column in range(start, end + 1):
                label = _norm(_value(labels_row, column))
                if label in {"my property", "my prop"}:
                    columns["property"] = column
                elif label in {"competitive set", "comp set"}:
                    columns["comp_set"] = column
                elif label.startswith("index "):
                    columns["index"] = column
            if {"property", "comp_set"} <= set(columns):
                result[metric] = columns
        if set(result) == {"occupancy", "adr", "revpar"}:
            return result
    raise StrWorkbookError("could not locate the metric columns")


def _parse_day_type(rows: list[list[Any]], sheet_name: str, report_month: date) -> dict[str, Any]:
    metric_columns = _metric_columns(rows)
    records = []
    range_start = None
    range_end = None
    current_scope = None
    for row_number, row in enumerate(rows, start=1):
        scope = _norm(_value(row, 2))
        if scope in {"weekday", "weekend"}:
            current_scope = scope
            range_start = row_number if range_start is None else min(range_start, row_number)
        if current_scope and _norm(_value(row, 3)) == "running 12 month":
            record = {"day_type": current_scope, **_running_12_period(report_month.year, report_month)}
            record["metrics"] = {
                metric: _metric_point(
                    _value(row, columns["property"]),
                    _value(row, columns["comp_set"]),
                    _value(row, columns.get("index", 0)),
                )
                for metric, columns in metric_columns.items()
            }
            records.append(record)
            range_end = row_number
            current_scope = None
    if {record["day_type"] for record in records} != {"weekday", "weekend"}:
        raise StrWorkbookError(f"{sheet_name} is missing running-12 weekday or weekend data")
    max_column = max(column for columns in metric_columns.values() for column in columns.values())
    return {
        "rows": records,
        "locator": _range(sheet_name, range_start or 1, 2, range_end or 1, max_column),
    }


def _segment_columns(rows: list[list[Any]]) -> tuple[int, dict[str, dict[str, int]]]:
    for row_number, row in enumerate(rows[:20], start=1):
        starts = [
            (column, _norm(value))
            for column, value in enumerate(row, start=1)
            if _norm(value) in {"transient", "group", "contract"}
        ]
        if {label for _, label in starts} != {"transient", "group", "contract"}:
            continue
        starts.sort()
        # STR segmentation sheets repeat the three segment headers in an adjacent
        # Percent Change block. Keep the first complete block, which contains the
        # actual values, instead of letting the repeated headers overwrite it.
        actual_starts = []
        seen = set()
        for start, segment in starts:
            if segment in seen:
                break
            actual_starts.append((start, segment))
            seen.add(segment)
        starts = actual_starts
        if {label for _, label in starts} != {"transient", "group", "contract"}:
            continue
        labels_row = rows[row_number] if row_number < len(rows) else []
        result: dict[str, dict[str, int]] = {}
        for index, (start, segment) in enumerate(starts):
            end = starts[index + 1][0] - 1 if index + 1 < len(starts) else min(len(labels_row), start + 3)
            columns: dict[str, int] = {}
            for column in range(start, end + 1):
                label = _norm(_value(labels_row, column))
                if label in {"my property", "my prop"}:
                    columns["property"] = column
                elif label in {"competitive set", "comp set"}:
                    columns["comp_set"] = column
            if {"property", "comp_set"} <= set(columns):
                result[segment] = columns
        if set(result) == {"transient", "group", "contract"}:
            return row_number, result
    raise StrWorkbookError("could not locate transient, group, and contract columns")


def _parse_segmentation(
    sheets: dict[str, list[list[Any]]],
    suffix: str,
    report_month: date,
) -> dict[str, Any]:
    combined: dict[tuple[int, str], dict[str, Any]] = defaultdict(dict)
    locators = []
    years_seen: set[int] = set()
    for base_name, metric in (
        ("Segmentation Occ", "occupancy"),
        ("Segmentation ADR", "adr"),
        ("Segmentation RevPAR", "revpar"),
    ):
        sheet_name = base_name + (f"_{suffix}" if suffix else "")
        if sheet_name not in sheets:
            raise StrWorkbookError(f"missing worksheet {sheet_name}")
        rows = sheets[sheet_name]
        _, columns_by_segment = _segment_columns(rows)
        running_row = next(
            (
                row_number
                for row_number, row in enumerate(rows, start=1)
                if any(_norm(value) == "running 12 month" for value in row[:8])
            ),
            None,
        )
        if running_row is None:
            raise StrWorkbookError(f"{sheet_name} is missing the Running 12 Month section")
        extracted_rows = []
        for row_number in range(running_row + 1, min(len(rows), running_row + 8) + 1):
            row = rows[row_number - 1]
            year = _integer(_value(row, 2))
            if year is None or not 1900 <= year <= 2200:
                if extracted_rows:
                    break
                continue
            extracted_rows.append(row_number)
            years_seen.add(year)
            for segment, columns in columns_by_segment.items():
                point = _metric_point(
                    _value(row, columns["property"]),
                    _value(row, columns["comp_set"]),
                )
                _validate_actual_metric_point(
                    point,
                    metric,
                    f"{sheet_name} Running 12 Month {year} {segment}",
                )
                combined[(year, segment)][metric] = point
            if len(extracted_rows) == 3:
                break
        if len(extracted_rows) != 3:
            raise StrWorkbookError(f"{sheet_name} does not contain three Running 12 Month years")
        max_column = max(column for columns in columns_by_segment.values() for column in columns.values())
        locators.append(_range(sheet_name, running_row, 2, extracted_rows[-1], max_column))

    rows_out = []
    for year in sorted(years_seen):
        for segment in ("transient", "group", "contract"):
            metrics = combined.get((year, segment), {})
            if set(metrics) != {"occupancy", "adr", "revpar"}:
                raise StrWorkbookError(f"segmentation data is incomplete for {year} {segment}")
            row = {"segment": segment, **_running_12_period(year, report_month), "metrics": metrics}
            rows_out.append(row)
    return {"rows": rows_out, "locator": "; ".join(locators)}


def _related_sheet(sheets: dict[str, list[list[Any]]], base_name: str, suffix: str) -> tuple[str, list[list[Any]]]:
    desired = base_name + (f"_{suffix}" if suffix else "")
    for sheet_name, rows in sheets.items():
        if sheet_name.casefold() == desired.casefold():
            return sheet_name, rows
    raise StrWorkbookError(f"missing worksheet {desired}")


def _evidence_record(
    ref: str,
    source_id: str,
    locator: str,
    subject: str,
    statement: str,
    excerpt: str,
    attributes: dict[str, Any],
) -> dict[str, Any]:
    return {
        "ref": ref,
        "source_id": source_id,
        "locator": {"kind": "workbook_range", "value": locator},
        "excerpt": excerpt,
        "extraction_method": "deterministic",
        "data": {
            "information_primitive": {
                "type": "property.str",
                "subject": subject,
                "statement": statement,
                "attributes": attributes,
            }
        },
    }


def _parse_family(
    sheets: dict[str, list[list[Any]]],
    comp_sheet_name: str,
    suffix: str,
    source_id: str,
    weekend_definition: str | None,
) -> dict[str, Any]:
    comp_rows = sheets[comp_sheet_name]
    report_month = _parse_report_month(comp_rows)
    report_month_text = report_month.strftime("%Y-%m")
    created_date = _parse_created_date(comp_rows)
    currency = _parse_currency(comp_rows)
    if not currency:
        for sheet_name, rows in sheets.items():
            if sheet_name.casefold().startswith("table of contents"):
                currency = _parse_currency(rows)
                if currency:
                    break
    title = _title(comp_rows)
    comp_set_key, comp_set_label, comp_set_number = _comp_set_identity(title, suffix)
    issues: list[str] = []

    hotels = None
    response_locator = None
    response_sheet_name = None
    try:
        response_sheet_name, response_rows = _related_sheet(sheets, "Response", suffix)
        hotels, response_locator = _find_response_hotels(response_rows)
        if response_sheet_name != "Response" and response_locator:
            response_locator = response_locator.replace("'Response'!", f"'{response_sheet_name}'!")
    except StrWorkbookError as exc:
        issues.append(str(exc))

    subject = hotels[0]["name"] if hotels else _subject_from_header(comp_rows)
    if not subject:
        raise StrWorkbookError(f"{comp_sheet_name} does not identify the subject hotel")
    subject_str_id = hotels[0].get("str_id") if hotels else None
    logical_label = "performance-set" if comp_set_key in {"competitive-set", "performance-set"} else comp_set_key
    logical_key = f"{subject_str_id or subject.casefold()}::{logical_label}"
    comp_set_version_id, comp_set_version_basis = _comp_set_version_identity(
        logical_label,
        subject,
        subject_str_id,
        hotels,
        source_id,
    )
    common = {
        "comp_set_key": comp_set_key,
        "comp_set_label": comp_set_label,
        "comp_set_version_id": comp_set_version_id,
        "comp_set_version_basis": comp_set_version_basis,
        "report_month": report_month_text,
        "report_period_end": _month_end(report_month).isoformat(),
        "penetration_definition": "property performance / comp-set performance * 100",
    }
    if comp_set_number is not None:
        common["comp_set_number"] = comp_set_number
    if subject_str_id:
        common["subject_str_id"] = subject_str_id
    if created_date:
        common["report_created_date"] = created_date
    if currency:
        common["currency"] = currency

    evidence: list[dict[str, Any]] = []
    ref_prefix = f"str-{comp_set_key}-{suffix or 'primary'}"
    if hotels and response_locator:
        evidence.append(
            _evidence_record(
                f"{ref_prefix}-membership",
                source_id,
                response_locator,
                subject,
                f"{comp_set_label} identifies the subject hotel and {len(hotels) - 1} competitive hotels with room counts and source-reported opening months.",
                f"{title}; response table for {subject} and {len(hotels) - 1} competitive hotels.",
                {**common, "observation_kind": "comp_set_membership", "hotels": hotels},
            )
        )

    comp = None
    try:
        comp = _parse_comp(comp_rows, comp_sheet_name, report_month)
    except StrWorkbookError as exc:
        issues.append(str(exc))
    if comp:
        units = {"occupancy": "percent", "adr": currency or "source_currency", "revpar": currency or "source_currency"}
        evidence.extend(
            [
                _evidence_record(
                    f"{ref_prefix}-monthly",
                    source_id,
                    comp["locators"]["monthly"],
                    subject,
                    f"{comp_set_label} reports 18 months of subject and comp-set occupancy, ADR, RevPAR, and penetration ending {report_month_text}.",
                    f"{title}; 18 monthly periods ending {report_month_text}.",
                    {**common, "observation_kind": "monthly_performance", "value_kind": "actual", "units": units, "rows": comp["monthly"]},
                ),
                _evidence_record(
                    f"{ref_prefix}-ytd",
                    source_id,
                    comp["locators"]["ytd"],
                    subject,
                    f"{comp_set_label} reports three years of YTD subject and comp-set occupancy, ADR, RevPAR, and penetration through the report month.",
                    f"{title}; three YTD periods ending in {report_month.strftime('%B')} of each reported year.",
                    {**common, "observation_kind": "ytd_performance", "value_kind": "actual", "units": units, "rows": comp["ytd"]},
                ),
                _evidence_record(
                    f"{ref_prefix}-running-12",
                    source_id,
                    comp["locators"]["running_12"],
                    subject,
                    f"{comp_set_label} reports three running-12-month subject and comp-set occupancy, ADR, RevPAR, and penetration periods.",
                    f"{title}; three Running 12 Month periods ending in {report_month.strftime('%B')} of each reported year.",
                    {**common, "observation_kind": "running_12_performance", "value_kind": "actual", "units": units, "rows": comp["running_12"]},
                ),
            ]
        )

    try:
        dow_sheet_name, dow_rows = _related_sheet(sheets, "Day of Week", suffix)
        dow = _parse_day_type(dow_rows, dow_sheet_name, report_month)
        day_attributes = {
            **common,
            "observation_kind": "day_type_running_12",
            "value_kind": "actual",
            "units": {"occupancy": "percent", "adr": currency or "source_currency", "revpar": currency or "source_currency"},
            "rows": dow["rows"],
        }
        if weekend_definition:
            day_attributes["weekend_definition"] = weekend_definition
        evidence.append(
            _evidence_record(
                f"{ref_prefix}-day-type-running-12",
                source_id,
                dow["locator"],
                subject,
                f"{comp_set_label} reports running-12-month weekday and weekend occupancy, ADR, RevPAR, and penetration through {report_month_text}.",
                f"{dow_sheet_name}; Running 12 Month weekday and weekend rows ending {report_month_text}.",
                day_attributes,
            )
        )
    except StrWorkbookError as exc:
        issues.append(str(exc))

    try:
        segmentation = _parse_segmentation(sheets, suffix, report_month)
        evidence.append(
            _evidence_record(
                f"{ref_prefix}-segment-running-12",
                source_id,
                segmentation["locator"],
                subject,
                f"{comp_set_label} reports three running-12-month years of transient, group, and contract occupancy, ADR, RevPAR, and calculated penetration.",
                f"Segmentation Running 12 Month tables for three years ending in {report_month.strftime('%B')}.",
                {
                    **common,
                    "observation_kind": "segment_running_12",
                    "value_kind": "actual",
                    "units": {"occupancy": "percent", "adr": currency or "source_currency", "revpar": currency or "source_currency"},
                    "penetration_index_method": "calculated as property / comp_set * 100 when the comp-set value is nonzero",
                    "rows": segmentation["rows"],
                },
            )
        )
    except StrWorkbookError as exc:
        issues.append(str(exc))

    evidence_kinds = {
        item["data"]["information_primitive"]["attributes"]["observation_kind"] for item in evidence
    }
    return {
        "source_id": source_id,
        "subject": subject,
        "subject_str_id": subject_str_id,
        "comp_set_key": comp_set_key,
        "comp_set_label": comp_set_label,
        "comp_set_version_id": comp_set_version_id,
        "comp_set_version_basis": comp_set_version_basis,
        "logical_key": logical_key,
        "report_month": report_month_text,
        "evidence": evidence,
        "evidence_kinds": sorted(evidence_kinds),
        "issues": issues,
        "status": "complete" if evidence_kinds == REQUIRED_EVIDENCE_KINDS else "partial",
    }


def parse_str_workbook(path: Path, source_id: str) -> list[dict[str, Any]]:
    sheets = load_workbook_rows(path)
    weekend_definition = _parse_weekend_definition(sheets)
    comp_families = []
    for sheet_name in sheets:
        match = re.fullmatch(r"Comp(?:_(\d+))?", sheet_name, re.IGNORECASE)
        if match:
            comp_families.append((match.group(1) or "", sheet_name))
    if not comp_families:
        raise StrWorkbookError("workbook has no Comp sheet family")
    comp_families.sort(key=lambda item: int(item[0]) if item[0] else 0)
    return [
        _parse_family(sheets, sheet_name, suffix, source_id, weekend_definition)
        for suffix, sheet_name in comp_families
    ]


def _candidate_sources(
    sources: Iterable[dict[str, Any]], source_ids: set[str] | None
) -> list[dict[str, Any]]:
    candidates = []
    for source in sources:
        if source_ids is not None and source.get("source_id") not in source_ids:
            continue
        location = source.get("location", {})
        extension = Path(str(location.get("value", ""))).suffix.casefold()
        if (
            source.get("document_type") == "performance.str_report"
            and source.get("source_kind") == "local_file"
            and location.get("kind") == "relative_path"
            and source.get("availability", "available") == "available"
            and extension in SUPPORTED_EXTENSIONS
        ):
            candidates.append(source)
    if source_ids is not None:
        missing = source_ids - {source.get("source_id") for source in candidates}
        if missing:
            raise StrWorkbookError(f"requested source IDs are not available STR workbooks: {sorted(missing)}")
    return candidates


def extract_str_findings(
    deal_room: Path,
    sources: list[dict[str, Any]],
    *,
    source_ids: set[str] | None = None,
    entity_name: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    root = deal_room.resolve()
    candidates = _candidate_sources(sources, source_ids)
    parsed: list[dict[str, Any]] = []
    skipped_sources: list[dict[str, str]] = []
    source_paths: dict[str, str] = {}
    for source in candidates:
        source_id = source["source_id"]
        relative_path = source["location"]["value"]
        path = (root / Path(relative_path)).resolve()
        try:
            path.relative_to(root)
        except ValueError:
            skipped_sources.append({"source_id": source_id, "path": relative_path, "reason": "path leaves deal-room root"})
            continue
        if not path.is_file():
            skipped_sources.append({"source_id": source_id, "path": relative_path, "reason": "file not found"})
            continue
        source_paths[source_id] = relative_path
        try:
            parsed.extend(parse_str_workbook(path, source_id))
        except (OSError, StrWorkbookError) as exc:
            skipped_sources.append({"source_id": source_id, "path": relative_path, "reason": str(exc)})

    selected = parsed

    canonical_entity_name = entity_name.strip() if entity_name else None
    bundles_by_source_subject: dict[tuple[str, str], dict[str, Any]] = {}
    for item in selected:
        resolved_subject = canonical_entity_name or item["subject"]
        for evidence in item["evidence"]:
            primitive = evidence["data"]["information_primitive"]
            primitive["subject"] = resolved_subject
            primitive["attributes"]["source_reported_subject"] = item["subject"]
            primitive["attributes"]["entity_name_resolution"] = (
                "model_supplied" if canonical_entity_name else "source_reported_preview_only"
            )
        key = (item["source_id"], resolved_subject)
        bundle = bundles_by_source_subject.setdefault(
            key,
            {
                "producer": dict(PRODUCER),
                "entity": {"entity_type": "hotel", "name": resolved_subject},
                "evidence": [],
            },
        )
        bundle["evidence"].extend(item["evidence"])

    comp_sets = []
    for item in sorted(selected, key=lambda value: (value["subject"], value["comp_set_label"])):
        resolved_subject = canonical_entity_name or item["subject"]
        monthly_attributes = next(
            (
                evidence["data"]["information_primitive"]["attributes"]
                for evidence in item["evidence"]
                if evidence["data"]["information_primitive"]["attributes"].get("observation_kind")
                == "monthly_performance"
            ),
            None,
        )
        monthly_rows = monthly_attributes.get("rows", []) if monthly_attributes else []
        comp_sets.append(
            {
                "source_id": item["source_id"],
                "path": source_paths.get(item["source_id"]),
                "subject": resolved_subject,
                "source_reported_subject": item["subject"],
                "subject_str_id": item["subject_str_id"],
                "comp_set_key": item["comp_set_key"],
                "comp_set_label": item["comp_set_label"],
                "comp_set_version_id": item["comp_set_version_id"],
                "comp_set_version_basis": item["comp_set_version_basis"],
                "report_month": item["report_month"],
                "monthly_period_start": monthly_rows[0]["period"] if monthly_rows else None,
                "monthly_period_end": monthly_rows[-1]["period"] if monthly_rows else None,
                "status": item["status"],
                "coverage": {
                    kind: "found" if kind in item["evidence_kinds"] else "not_found"
                    for kind in sorted(REQUIRED_EVIDENCE_KINDS)
                },
                "issues": item["issues"],
                "evidence_count": len(item["evidence"]),
            }
        )

    status = "not_available"
    if comp_sets:
        status = "complete" if all(item["status"] == "complete" for item in comp_sets) else "partial"
    coverage = {
        "producer": dict(PRODUCER),
        "status": status,
        "entity_name_resolution": "model_supplied" if canonical_entity_name else "source_reported_preview_only",
        "selection_policy": "requested STR reports" if source_ids is not None else "all available STR reports",
        "sources_considered": len(candidates),
        "comp_sets_selected": len(comp_sets),
        "evidence_prepared": sum(item["evidence_count"] for item in comp_sets),
        "facts_prepared": 0,
        "comp_sets": comp_sets,
        "older_reports_not_selected": [],
        "skipped_sources": skipped_sources,
    }
    return list(bundles_by_source_subject.values()), coverage
